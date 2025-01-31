from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side, Alignment, Color
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.worksheet.worksheet import Worksheet
from Listpy import Convertir_Lista_De_Tuplas_A_Lista_De_Listas
import pandas as pd
from typing import Optional, List
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side

def Obtener_Hojas_Y_Nombres(Libro):

    """
    Obtiene los nombres de las hojas y los objetos de hoja correspondientes 
    de un libro de trabajo.

    Parametros:
    - Libro: El objeto libro de trabajo del cual extraer nombres y hojas.

    Retorna:
    - tuple: Una tupla que contiene una lista de nombres de hojas y una 
      lista de hojas correspondientes.

    """

    Nombres_Hojas = Libro.sheetnames 
    Hojas = []
    for Hoja in Nombres_Hojas:
        Hoja_Actual = Libro[Hoja]
        Hojas.append(Hoja_Actual)
    return Nombres_Hojas, Hojas

def Obtener_Valores_Rango(Celda1: str, Celda2: str, Hoja) -> list:

    """
    Obtiene valores de un rango específico en una hoja de trabajo.

    Parametros:
    - Celda1 (str): La celda inicial (ej: 'A1').
    - Celda2 (str): La celda final (ej: 'B2').
    - Hoja: El objeto hoja de trabajo de la cual extraer valores.

    Retorna:
    - list: Una lista de listas conteniendo los valores del rango 
      especificado.

    """

    Rango = Hoja[Celda1:Celda2]
    Rango = Convertir_Lista_De_Tuplas_A_Lista_De_Listas(Rango)
    Valores_Rango = []

    for Fila in Rango:
        Valores_Fila = []
        for Celda in Fila:
            Valores_Fila.append(Celda.value)
        Valores_Rango.append(Valores_Fila)
    return Valores_Rango

def Verificar_Valor_Celda(Valor, Hoja, Columna: int, Fila: int) -> bool:

    """
    Verifica si el valor de una celda coincide con el valor especificado.

    Parametros:
    - Valor: El valor contra el cual comparar.
    - Hoja: El objeto hoja de trabajo que contiene la celda.
    - Columna (int): El número de columna de la celda (índice base 1).
    - Fila (int): El número de fila de la celda (índice base 1).

    Retorna:
    - bool: True si el valor de la celda coincide, False en caso contrario.

    """
    
    return Valor == Hoja.cell(row=Fila, column=Columna).value

def Obtener_Valor_Celda(Valor, Hoja, Celda: Optional[str] = None, 
                       Columna: Optional[int] = None, 
                       Fila: Optional[int] = None):
    
    """
    Obtiene el valor de una celda en una hoja de trabajo.

    Parametros:
    - Valor: El valor contra el cual comparar.
    - Hoja: El objeto hoja de trabajo que contiene la celda.
    - Celda (str, opcional): La referencia de celda (ej: 'A1').
    - Columna (int, opcional): El número de columna (índice base 1).
    - Fila (int, opcional): El número de fila (índice base 1).

    Retorna:
    - El valor de la celda especificada.

    Eleva:
    - KeyError: Si no se proporciona una referencia de celda o columna/fila.

    """

    if Celda is None and Columna is None and Fila is None:
        raise KeyError(
            'Debe proporcionar una referencia de celda (ej: "A1") o ambos '
            'números de columna y fila.'
        )
    
    if Celda:
        return Valor == Hoja[Celda].value
    
    return Hoja.cell(row=Fila, column=Columna).value

def Obtener_Largo_Y_Ancho(Hoja) -> tuple[int, int]:

    """
    Obtiene las dimensiones (largo y ancho) de una hoja de trabajo.

    Parametros:
    - Hoja: El objeto hoja de trabajo a medir.

    Retorna:
    - tuple: Una tupla conteniendo el número de filas (largo) y 
      columnas (ancho).

    """

    Largo = Hoja.max_row  
    Ancho = Hoja.max_column
    return Largo, Ancho

def Combinar_Hojas(Libro, Nombre):

    """
    Combina todas las hojas de un libro en una nueva hoja.

    Parametros:
    - Libro: El objeto libro de trabajo que contiene las hojas a combinar.
    - Nombre: El nombre para la nueva hoja combinada.

    """

    Nuevo_Doc = Workbook()
    Nueva_Hoja = Nuevo_Doc.create_sheet(Nombre)
    Nombres_Hojas, Hojas = Obtener_Hojas_Y_Nombres(Libro)
    Largo_Total = 1

    for Hoja in Hojas:
        Largo_Hoja, Ancho_Hoja = Obtener_Largo_Y_Ancho(Hoja)    
        for Fila in range(Largo_Total, Largo_Total + Largo_Hoja):
            for Columna in range(1, Ancho_Hoja + 1):
                Celda = Hoja.cell(row=Fila - Largo_Total + 1, 
                                column=Columna).value
                Nueva_Hoja.cell(row=Fila, column=Columna, value=Celda)
        Largo_Total += Largo_Hoja

    Nuevo_Doc.save(Nombre)

def Formatear_Libro(Ruta: str, Nombre_Hoja: Optional[str] = None):

    """
    Formatea un libro aplicando filtros, alineación y bordes.

    Parametros:
    - Ruta (str): La ruta al archivo del libro.
    - Nombre_Hoja (str, opcional): El nombre de la hoja a formatear. 
      Por defecto es la hoja activa.

    """

    Libro = load_workbook(Ruta)
    
    if Nombre_Hoja is None:
        Hoja = Libro.active  
    else:
        if Nombre_Hoja not in Libro.sheetnames:
            raise ValueError(
                f"La hoja '{Nombre_Hoja}' no existe en el libro."
            )
        Hoja = Libro[Nombre_Hoja]

    # Aplicar autofiltro.
    if Hoja is not None:
        Hoja.auto_filter.ref = Hoja.dimensions
    else:
        raise ValueError("La hoja especificada no existe.")

    Formato_Alineacion = Alignment(horizontal='center', vertical='center')
    Formato_Borde = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'),
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    for Columna in range(1, Hoja.max_column + 1):
        for Fila in range(1, Hoja.max_row + 1):
            Celda = Hoja.cell(row=Fila, column=Columna)
            Celda.alignment = Formato_Alineacion
            Celda.border = Formato_Borde

    Libro.save(Ruta)

def Convertir_DataFrame_A_Excel(df, Ruta: str, Nombre_Archivo: str):

    """
    Convierte un DataFrame a un archivo Excel con formato.

    Parametros:
    - df: El DataFrame a convertir.
    - Ruta (str): La ruta del directorio donde se guardará el archivo.
    - Nombre_Archivo (str): El nombre del archivo Excel.

    """

    Ruta_Completa = f"{Ruta}{Nombre_Archivo}"
    Formatear_Libro(Ruta_Completa)

def Inmovilizar_Paneles(Ruta: str, Celda: str, 
                       Nombre_Hoja: Optional[str] = None):

    """
    Congela los paneles en una hoja de Excel especificada.

    Parametros:
    - Ruta (str): La ruta al archivo del libro.
    - Celda (str): La celda en la que congelar los paneles (ej: 'A1').
    - Nombre_Hoja (str, opcional): El nombre de la hoja a modificar. 
      Por defecto es la hoja activa.

    """

    Libro = load_workbook(Ruta)
    
    if Nombre_Hoja is None:
        Hoja = Libro.active
    else:
        if Nombre_Hoja not in Libro.sheetnames:
            raise ValueError(f"La hoja '{Nombre_Hoja}' no existe.")
        Hoja = Libro[Nombre_Hoja]

    if Hoja is not None:
        Hoja.freeze_panes = Celda
        Libro.save(Ruta)
    else:
        raise ValueError("La hoja especificada no existe.")

def Reemplazar_Caracter_En_Rango_Excel(
    Ruta: str, 
    Caracter_Viejo: str, 
    Caracter_Nuevo: str, 
    Col_Min: int, 
    Fila_Min: int, 
    Col_Max: Optional[int] = None, 
    Fila_Max: Optional[int] = None, 
    Nombre_Hoja: Optional[str] = None
):
    
    """
    Reemplaza caracteres en un rango específico de un libro Excel.

    Parametros:
    - Ruta (str): La ruta al archivo del libro.
    - Caracter_Viejo (str): El carácter a reemplazar.
    - Caracter_Nuevo (str): El carácter a usar como reemplazo.
    - Col_Min (int): El índice de columna mínimo (base 1).
    - Fila_Min (int): El índice de fila mínimo (base 1).
    - Col_Max (int, opcional): El índice de columna máximo (base 1).
    - Fila_Max (int, opcional): El índice de fila máximo (base 1).
    - Nombre_Hoja (str, opcional): El nombre de la hoja a modificar.
      Por defecto es la hoja activa.

    """

    Libro = load_workbook(Ruta)
    
    if Nombre_Hoja is None:
        Hoja = Libro.active
    else:
        Hoja = Libro[Nombre_Hoja]
    
    if Hoja is None:
        raise ValueError("La hoja especificada no existe.")
    
    if Col_Max is None:
        Col_Max = Hoja.max_column
    if Fila_Max is None:
        Fila_Max = Hoja.max_row
    
    for Fila in Hoja.iter_rows(min_row=Fila_Min, min_col=Col_Min, 
                             max_row=Fila_Max, max_col=Col_Max):  
        for Celda in Fila:
            Contenido = Celda.value
            if Contenido:
                if isinstance(Contenido, str):
                    Celda.value = Contenido.replace(
                        Caracter_Viejo, 
                        Caracter_Nuevo
                    )
    
    Libro.save(Ruta)

def Eliminar_Columnas_Excel(
    Ruta: str, 
    Columnas: list[str], 
    Nombre_Hoja: Optional[str] = None
):

    """
    Elimina columnas específicas de una hoja de Excel.

    Parametros:
    - Ruta (str): La ruta al archivo del libro.
    - Columnas (list[str]): Lista de nombres de columnas a eliminar.
    - Nombre_Hoja (str, opcional): El nombre de la hoja a modificar.
      Por defecto es la hoja activa.

    """

    Libro = load_workbook(Ruta)
    
    if Nombre_Hoja is None:
        Hoja = Libro.active
    else:
        Hoja = Libro[Nombre_Hoja]
    
    if Hoja is None:
        raise ValueError("La hoja especificada no existe.")
    
    Fila1 = Hoja[1]
    
    Mapeo = {Celda.value: Celda.column for Celda in Fila1}
    
    Indice_Columnas = [
        Mapeo[i] for i in Columnas if i in Mapeo
    ]
     
    for i in sorted(Indice_Columnas, reverse=True):
        Hoja.delete_cols(i)
    
    Libro.save(Ruta)

def Ajustar_Ancho_Columna(
    Ruta: str, 
    Columna_Min: int, 
    Columna_Max: int, 
    Ancho: float, 
    Nombre_Hoja: Optional[str] = None
):
    
    """
    Ajusta el ancho de columnas específicas en una hoja de trabajo.

    Parametros:
    - Ruta (str): La ruta al archivo del libro.
    - Columna_Min (int): El índice de columna inicial (base 1).
    - Columna_Max (int): El índice de columna final (base 1).
    - Ancho (float): El ancho a establecer para las columnas.
    - Nombre_Hoja (str, opcional): El nombre de la hoja a modificar.
      Por defecto es la hoja activa.

    """

    Libro = load_workbook(Ruta)
    
    if Nombre_Hoja is None:
        Hoja = Libro.active
    else:
        Hoja = Libro[Nombre_Hoja]
        
    if Hoja is None:
        raise ValueError("La hoja especificada no existe.")
        
    for Columna in range(Columna_Min, Columna_Max + 1):
        Hoja.column_dimensions[get_column_letter(Columna)].width = Ancho
    
    Libro.save(Ruta)

def Crear_Excel_Con_Multiple_DataFrames(
    Ruta: str, 
    Lista_Df: list[pd.DataFrame], 
    Lista_Nombres_Hojas: Optional[List[str]] = None
):
    
    """
    Crea un libro Excel con múltiples DataFrames en hojas separadas.

    Parametros:
    - Ruta (str): La ruta donde se guardará el libro.
    - Lista_Df (list[pd.DataFrame]): Lista de DataFrames a guardar.
    - Lista_Nombres_Hojas (list[str], opcional): Lista de nombres para las 
      hojas. Si no se proporciona, por defecto será "Hoja1", "Hoja2", etc.
    
    Eleva:
    - ValueError: Si las longitudes de Lista_Df y Lista_Nombres_Hojas no 
      coinciden.

    """

    if Lista_Nombres_Hojas is None:
        Lista_Nombres_Hojas = [
            f'Hoja{i+1}' for i in range(len(Lista_Df))
        ]

    elif len(Lista_Df) != len(Lista_Nombres_Hojas):
        raise ValueError(
            "La lista de nombres de hojas debe coincidir con la longitud "
            "de la lista de DataFrames."
        )

    with pd.ExcelWriter(Ruta) as Escritor:
        for Df, Nombre_Hoja in zip(Lista_Df, Lista_Nombres_Hojas):
            Df.to_excel(Escritor, sheet_name = Nombre_Hoja, index = False)
