from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side, Alignment, Color
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.worksheet.worksheet import Worksheet
from Listio import *
import pandas as pd
from typing import Optional, List
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side

def Get_Sheets_And_Sheetnames(Book):

    """
    Get the sheet names and corresponding sheet objects from a workbook.

    Parameters:
    - Book: The workbook object to extract sheet names and sheets from.

    Returns:
    - tuple: A tuple containing a list of sheet names and a list of 
      corresponding sheets.

    """

    Sheet_Names = Book.sheetnames 
    Sheets = []
    for Sheet in Sheet_Names:  # Fixed to iterate over Sheet_Names
        Current_Sheet = Book[Sheet]
        Sheets.append(Current_Sheet)
    return Sheet_Names, Sheets

def Get_Range_Values(Cell1: str, Cell2: str, Sheet) -> list:

    """
    Get values from a specified range in a worksheet.

    Parameters:
    - Cell1 (str): The starting cell (e.g., 'A1').
    - Cell2 (str): The ending cell (e.g., 'B2').
    - Sheet: The worksheet object from which to extract values.

    Returns:
    - list: A list of lists containing the values from the specified range.

    """

    Range = Sheet[Cell1:Cell2]
    Range = Convert_List_Of_Tuples_To_List_Of_List(Range)
    Range_Values = []

    for Row in Range:
        Row_Values = []
        for Cell in Row:
            Row_Values.append(Cell.value)
        Range_Values.append(Row_Values)
    return Range_Values

def Verify_Cell_Value(Value, Sheet, Column: int, Row: int) -> bool:

    """
    Verify if a cell's value matches the specified value.

    Parameters:
    - Value: The value to compare against.
    - Sheet: The worksheet object containing the cell.
    - Column (int): The column number of the cell (1-based index).
    - Row (int): The row number of the cell (1-based index).

    Returns:
    - bool: True if the cell's value matches, False otherwise.

    """
    
    return Value == Sheet.cell(row=Row, column=Column).value

def Get_Cell_Value(Value, Sheet, Cell: Optional[str] = None, Column: Optional[int] = None, 
                   Row: Optional[int] = None):
    
    """
    Get the value of a cell in a worksheet.

    Parameters:
    - Value: The value to compare against.
    - Sheet: The worksheet object containing the cell.
    - Cell (str, optional): The cell reference (e.g., 'A1').
    - Column (int, optional): The column number (1-based index).
    - Row (int, optional): The row number (1-based index).

    Returns:
    - The value of the specified cell.

    Raises:
    - KeyError: If no cell reference or column/row is provided.

    """

    if Cell is None and Column is None and Row is None:
        raise KeyError(
            'You must provide a cell reference (e.g., "A1") or both '
            'column and row numbers.'
        )
    
    if Cell:
        return Value == Sheet[Cell].value
    
    return Sheet.cell(row=Row, column=Column).value

def Get_Length_And_Width(Sheet) -> tuple[int, int]:

    """
    Get the dimensions (length and width) of a worksheet.

    Parameters:
    - Sheet: The worksheet object to measure.

    Returns:
    - tuple: A tuple containing the number of rows (length) and 
      columns (width).

    """

    Length = Sheet.max_row  
    Width = Sheet.max_column
    return Length, Width

def Merge_Sheets(Book, Name):

    """
    Merge all sheets in a workbook into a new sheet.

    Parameters:
    - Book: The workbook object containing the sheets to merge.
    - Name: The name for the new merged sheet.

    """

    New_Doc = Workbook()
    New_Sheet = New_Doc.create_sheet(Name)
    Sheet_Names, Sheets = Get_Sheets_And_Sheetnames(Book)
    Total_Length = 1

    for Sheet in Sheets:
        Sheet_Length, Sheet_Width = Get_Length_And_Width(Sheet)    
        for Row in range(Total_Length, Total_Length + Sheet_Length):
            for Column in range(1, Sheet_Width + 1):  # Fixed to include last column
                Cell = Sheet.cell(row=Row - Total_Length + 1, column=Column).value
                New_Sheet.cell(row=Row, column=Column, value=Cell)
        Total_Length += Sheet_Length

    New_Doc.save(Name)  # Saving the new workbook

def Formating_Book(Path: str, Sheet_Name: Optional[str] = None):

    """
    Format a workbook by applying filters, alignment, and borders.

    Parameters:
    - Path (str): The path to the workbook file.
    - Sheet_Name (str, optional): The name of the sheet to format. 
      Defaults to the active sheet.

    """

    Book = load_workbook(Path)
    
    if Sheet_Name is None:
        Sheet = Book.active  
    else:
        if Sheet_Name not in Book.sheetnames:
            raise ValueError(f"Sheet '{Sheet_Name}' does not exist in the workbook.")
        Sheet = Book[Sheet_Name]

    # Apply autofilter.
    Sheet.auto_filter.ref = Sheet.dimensions

    Align_Format = Alignment(horizontal='center', vertical='center')
    Border_Format = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
    
    for Column in range(1, Sheet.max_column + 1):
        for Row in range(1, Sheet.max_row + 1):
            Cell = Sheet.cell(row=Row, column=Column)
            Cell.alignment = Align_Format
            Cell.border = Border_Format

    Book.save(Path)

def Convert_DataFrame_To_Excel(df, Path: str, Filename: str):

    """
    Convert a DataFrame to an Excel file with formatting.

    Parameters:
    - df: The DataFrame to convert.
    - Path (str): The directory path where the file will be saved.
    - Filename (str): The name of the Excel file.

    """

    Full_Path = f"{Path}{Filename}"
    Formating_Book(Full_Path)

def Immobilize_Panes(Path: str, Cell: str, Sheet_Name: Optional[str] = None):

    """
    Freeze panes in a specified Excel worksheet.

    Parameters:
    - Path (str): The path to the workbook file.
    - Cell (str): The cell at which to freeze panes (e.g., 'A1').
    - Sheet_Name (str, optional): The name of the sheet to modify. 
      Defaults to the active sheet.

    """

    Book = load_workbook(Path)
    
    if Sheet_Name is None:
        Sheet = Book.active
    else:
        if Sheet_Name not in Book.sheetnames:
            raise ValueError(f"Sheet '{Sheet_Name}' does not exist.")
        Sheet = Book[Sheet_Name]

    Sheet.freeze_panes = Cell
    Book.save(Path)

def Replace_Character_In_Range_Of_Excel_Book(
    Path: str, Old_Character: str, New_Character: str, 
    Min_Col: int, Min_Row: int, Max_Col: Optional[int] = None, 
    Max_Row: Optional[int] = None, Sheet_Name: Optional[str] = None
):
    
    """
    Replace characters in a specified range of an Excel workbook.

    Parameters:
    - Path (str): The path to the workbook file.
    - Old_Character (str): The character to replace.
    - New_Character (str): The character to use as replacement.
    - Min_Col (int): The minimum column index (1-based).
    - Min_Row (int): The minimum row index (1-based).
    - Max_Col (int, optional): The maximum column index (1-based).
    - Max_Row (int, optional): The maximum row index (1-based).
    - Sheet_Name (str, optional): The name of the sheet to modify.
      Defaults to the active sheet.

    """

    Book = load_workbook(Path)
    
    if Sheet_Name is None:
        Sheet = Book.active
    else:
        Sheet = Book[Sheet_Name]
    
    if Max_Col is None:
        Max_Col = Sheet.max_column
    if Max_Row is None:
        Max_Row = Sheet.max_row
    
    for Row in Sheet.iter_rows(min_row=Min_Row, min_col=Min_Col, 
                               max_row=Max_Row, max_col=Max_Col):  
        for Cell in Row:
            Content = Cell.value
            if Content:
                Cell.value = Content.replace(Old_Character, New_Character)
    
    Book.save(Path)

def Delete_Columns_In_Excel_Book(Path: str, Columns: list[str], Sheet_Name: Optional[str] = None):

    """
    Delete specified columns from an Excel worksheet.

    Parameters:
    - Path (str): The path to the workbook file.
    - Columns (list[str]): List of column names to delete.
    - Sheet_Name (str, optional): The name of the sheet to modify.
      Defaults to the active sheet.

    """

    Book = load_workbook(Path)
    
    if Sheet_Name is None:
        Sheet = Book.active
    else:
        Sheet = Book[Sheet_Name]
    
    Row1 = Sheet[1]
    
    Mapping = {Cell.value: Cell.column for Cell in Row1}
    
    Column_Index = [Mapping[i] for i in Columns if i in Mapping]
     
    for i in sorted(Column_Index, reverse=True):
        Sheet.delete_cols(i)
    
    Book.save(Path)

def Adjust_Column_Width(Path: str, Min_Column: int, Max_Column: int, 
                         Width: float, Sheet_Name: Optional[str] = None):
    
    """
    Adjust the width of specified columns in a worksheet.

    Parameters:
    - Path (str): The path to the workbook file.
    - Min_Column (int): The starting column index (1-based).
    - Max_Column (int): The ending column index (1-based).
    - Width (float): The width to set for the columns.
    - Sheet_Name (str, optional): The name of the sheet to modify.
      Defaults to the active sheet.

    """

    Book = load_workbook(Path)
    
    if Sheet_Name is None:
        Sheet = Book.active
    else:
        Sheet = Book[Sheet_Name]
        
    for Column in range(Min_Column, Max_Column + 1):
        Sheet.column_dimensions[chr(64 + Column)].width = Width
    
    Book.save(Path)

def Create_Excel_With_Multiple_Data_Frames(Path: str, df_List: list[pd.DataFrame], 
                                            Sheet_Names_List: Optional[List[str]] = None):
    
    """
    Create an Excel workbook with multiple DataFrames in separate sheets.

    Parameters:
    - Path (str): The path where the workbook will be saved.
    - df_List (list[pd.DataFrame]): List of DataFrames to save.
    - Sheet_Names_List (list[str], optional): List of names for the sheets. 
      If not provided, defaults to "Sheet1", "Sheet2", etc.
    
    Raises:
    - ValueError: If the lengths of df_List and Sheet_Names_List do not match.

    """

    if Sheet_Names_List is None:
        Sheet_Names_List = [f'Sheet{i+1}' for i in range(len(df_List))]

    elif len(df_List) != len(Sheet_Names_List):
        raise ValueError("The list of sheet names must match the length of "
                         "the DataFrames list.")

    with pd.ExcelWriter(Path) as Writer:
        for df, Sheet_Name in zip(df_List, Sheet_Names_List):
            df.to_excel(Writer, sheet_name=Sheet_Name, index=False)
